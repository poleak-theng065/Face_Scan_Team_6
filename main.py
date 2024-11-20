import cv2
import numpy as np
import face_recognition
import os
from openpyxl import Workbook, load_workbook
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image
import uuid
from datetime import datetime

excel_file = "data.xlsx"
images = []
classNames = []
imagePaths = []

# Load data from Excel
wb = load_workbook(excel_file)
ws = wb.active

data_dict = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    name = row[0]
    info = {
        "age": row[1],
        "gender": row[2],
        "hometown": row[3],
        "address": row[4],
        "Image": row[5]  # Assuming the image path is in the 6th column
    }
    data_dict[name] = info
    # Load the image from the path in the Excel file
    curImg = cv2.imread(info["Image"])
    if curImg is not None:
        images.append(curImg)
        classNames.append(name)
        imagePaths.append(info["Image"])
    else:
        print(f"Warning: Failed to load image for {name} from path: {info['Image']}")

print("Class names:", classNames)
def findEncodings(images):
        encodeList = []
        for img in images:
            try:
                encodes = face_recognition.face_encodings(img)
                if encodes:  # Check if encodes are found to avoid empty lists
                    encodeList.append(encodes[0])
                else:
                    print("Warning: No encodings found for image.")
            except Exception as e:
                print(f"Error encoding image: {e}")
        return encodeList
encodeListKnown = findEncodings(images)
print('Encoding Complete. Number of encodings:', len(encodeListKnown))

# Function for check attandance 
def markAttendance(name):
    file_name = 'Attendance.xlsx'
    
    # Check if the Excel file exists, if not, create it and add the header
    if not os.path.isfile(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Attendance'
        sheet.append(['Name', 'Time'])  # Adding the header
        workbook.save(file_name)
    
    # Open the Excel file and read data
    workbook = load_workbook(file_name)
    sheet = workbook.active
    
    # Collect existing names to prevent duplicates
    nameList = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]
    
    if name not in nameList:
        now = datetime.now()
        dtString = now.strftime('%H:%M:%S')
        sheet.append([name, dtString])  # Append new name and time
        workbook.save(file_name)  # Save changes to the file
# Face Detection and Verification
def verifydata():
    cap = cv2.VideoCapture(0)

    while True:
        success, img = cap.read()
        if not success:
            print("Failed to grab frame from webcam.")
            break

        imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

        facesCurFrame = face_recognition.face_locations(imgS)
        encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

        for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
            matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
            
            name = "Unknown"  # Default to "Unknown" if no match is found

            if matches and any(matches):
                matchIndex = np.argmin(faceDis)
                if matches[matchIndex]:
                    name = classNames[matchIndex]  # Update name if a match is found

            # Draw rectangle around the face
            y1, x2, y2, x1 = faceLoc
            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
            cv2.rectangle(img, (x1, y1), (x2, y2), (255, 0, 0), 2)

            # Display name above the rectangle
            label_y = y1 - 10 if y1 - 10 > 10 else y1 + 20
            cv2.putText(img, name, (x1, label_y), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)

            # If the name is not "Unknown", display additional info card
            if name != "Unknown":
                # Fetch data
                info = data_dict.get(name, {})
                if info:
                    # Define card position above the rectangle
                    card_x1, card_y1 = x1, y1 - 130
                    card_x2, card_y2 = x2, y1 - 10

                    # Draw card background
                    cv2.rectangle(img, (card_x1, card_y1), (card_x2, card_y2), (0, 0, 255), cv2.FILLED)
                    
                    # Draw connecting line
                    cv2.line(img, (x1, y1), (card_x1, card_y2), (255, 255, 255), 2)

                    # Add text to the card
                    cv2.putText(img, f"Name: {name}", (card_x1 + 6, card_y1 + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
                    cv2.putText(img, f"Age: {info.get('age', 'N/A')}", (card_x1 + 6, card_y1 + 40), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
                    cv2.putText(img, f"Gender: {info.get('gender', 'N/A')}", (card_x1 + 6, card_y1 + 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
                    cv2.putText(img, f"Hometown: {info.get('hometown', 'N/A')}", (card_x1 + 6, card_y1 + 80), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
                    cv2.putText(img, f"Address: {info.get('address', 'N/A')}", (card_x1 + 6, card_y1 + 100), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
                    # Attandance checked 
                    markAttendance(name)
        cv2.imshow('Webcam', img)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()



# Function to save data to Excel
def save_data():
    name = name_entry.get()
    age = age_entry.get()
    gender = gender_var.get()
    hometown = hometown_entry.get()
    class_name = class_entry.get()
    image_path = image_label.cget("text")  # Get the image path from the label
    
    if name and age and gender and hometown and class_name and image_path != "No Image Selected":
        try:
            # Create image folder if it doesn't exist
            if not os.path.exists('image'):
                os.makedirs('image')

            # Generate a unique filename for the image
            image_filename = f"{uuid.uuid4().hex}.png"
            image_save_path = os.path.join('image', image_filename)

            # Open and save the image to the 'image' folder
            img = Image.open(image_path)
            img.save(image_save_path)

            # Load the existing Excel file or create a new one if it doesn't exist
            try:
                df = pd.read_excel('data.xlsx')
            except FileNotFoundError:
                df = pd.DataFrame(columns=['Name', 'Age', 'Gender', 'Hometown', 'Class', 'Image'])

            # Add new data to DataFrame
            new_data = pd.DataFrame([[name, age, gender, hometown, class_name, image_save_path]],
                                    columns=['Name', 'Age', 'Gender', 'Hometown', 'Class', 'Image'])
            df = pd.concat([df, new_data], ignore_index=True)

            # Save updated data to Excel
            df.to_excel('data.xlsx', index=False)

            messagebox.showinfo("Success", "Data has been saved successfully!")
            clear_entries()
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while saving data: {e}")
    else:
        messagebox.showwarning("Input Error", "Please fill out all fields and select an image.")

# Function to clear the input fields
def clear_entries():
    name_entry.delete(0, tk.END)
    age_entry.delete(0, tk.END)
    hometown_entry.delete(0, tk.END)
    class_entry.delete(0, tk.END)
    gender_var.set(None)
    image_label.config(text="No Image Selected")

# Function to select an image
def select_image():
    file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png;*.bmp")])
    if file_path:
        image_label.config(text=file_path)

# Create main window
window = tk.Tk()
window.title("Student Data Entry Form")
window.geometry("500x600")

# Header Frame (using grid instead of pack)
header_frame = tk.Frame(window, bg="#4CAF50", bd=0, relief="flat")
header_frame.grid(row=0, column=0, columnspan=3, pady=10, sticky="ew")

header_label = tk.Label(header_frame, text="Student Data Entry", font=("Helvetica", 18, "bold"))
header_label.grid(row=10, column=10, padx=150, pady=10)

# Create labels and entry fields (with modern look)
font_style = ("Arial", 12)

# Name Entry
name_label = tk.Label(window, text="Name:", font=font_style, bg="#f0f0f0")
name_label.grid(row=1, column=0, padx=20, pady=10, sticky="w")

global name_entry
name_entry = tk.Entry(window, font=font_style, bd=2, relief="solid")
name_entry.grid(row=1, column=1, padx=20, pady=10, sticky="w")

# Age Entry
age_label = tk.Label(window, text="Age:", font=font_style, bg="#f0f0f0")
age_label.grid(row=2, column=0, padx=20, pady=10, sticky="w")

global age_entry
age_entry = tk.Entry(window, font=font_style, bd=2, relief="solid")
age_entry.grid(row=2, column=1, padx=20, pady=10, sticky="w")

# Gender Options
gender_label = tk.Label(window, text="Gender:", font=font_style, bg="#f0f0f0")
gender_label.grid(row=3, column=0, padx=20, pady=10, sticky="w")

global gender_var
gender_var = tk.StringVar(value=None)
male_rb = tk.Radiobutton(window, text="Male", variable=gender_var, value="Male", font=font_style, bg="#f0f0f0")
male_rb.grid(row=3, column=1, padx=10, pady=5, sticky="w")

female_rb = tk.Radiobutton(window, text="Female", variable=gender_var, value="Female", font=font_style, bg="#f0f0f0")
female_rb.grid(row=3, column=2, padx=10, pady=5, sticky="w")

# Hometown Entry
hometown_label = tk.Label(window, text="Hometown:", font=font_style, bg="#f0f0f0")
hometown_label.grid(row=4, column=0, padx=20, pady=10, sticky="w")

global hometown_entry
hometown_entry = tk.Entry(window, font=font_style, bd=2, relief="solid")
hometown_entry.grid(row=4, column=1, padx=20, pady=10, sticky="w")

# Class Entry
class_label = tk.Label(window, text="Class:", font=font_style, bg="#f0f0f0")
class_label.grid(row=5, column=0, padx=20, pady=10, sticky="w")

global class_entry
class_entry = tk.Entry(window, font=font_style, bd=2, relief="solid")
class_entry.grid(row=5, column=1, padx=20, pady=10, sticky="w")

# Image selection
global image_label
image_label = tk.Label(window, text="No Image Selected", width=30, relief="solid", font=("Arial", 10), bg="#f0f0f0")
image_label.grid(row=6, column=0, columnspan=2, padx=20, pady=10, sticky="w")

select_image_button = tk.Button(window, text="Select Image", command=select_image, font=("Arial", 12), bg="#4CAF50", fg="white", relief="flat")
select_image_button.grid(row=6, column=2, padx=10, pady=10)

# Submit Button
submit_button = tk.Button(window, text="Submit", command=save_data, font=("Arial", 12, "bold"), bg="#4CAF50", fg="white", relief="flat")
submit_button.grid(row=7, column=0, columnspan=3, pady=20)
# Verify Data
verify_button = tk.Button(window, text="Verify Data", command=verifydata, font=("Arial", 12, "bold"), bg="#4CAF50", fg="white", relief="flat")
verify_button.grid(row=14, column=0, columnspan=6, pady=40)

# Run the application
window.mainloop()

