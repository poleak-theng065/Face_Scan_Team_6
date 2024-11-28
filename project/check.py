import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

def markAttendance(name, action="check-in"):
    file_name = 'Attendance.xlsx'

    # Create the file if it doesn't exist
    if not os.path.isfile(file_name):
        workbook = Workbook()
        workbook.active.append(['Name', 'Date', 'Check-In Time', 'Check-In Status', 'Check-Out Time', 'Check-Out Status'])
        workbook.save(file_name)

    workbook = load_workbook(file_name)
    sheet = workbook.active
    now = datetime.now()
    today = now.strftime('%Y-%m-%d')
    time_now = now.strftime('%H:%M:%S')

    # Define thresholds for late/early
    check_in_deadline = datetime.strptime('07:00:00', '%H:%M:%S').time()
    check_out_deadline = datetime.strptime('16:00:00', '%H:%M:%S').time()

    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == name and row[1].value == today:
            # Update check-out time
            if action == "check-out":
                row[4].value = time_now
                check_out_time = datetime.strptime(time_now, '%H:%M:%S').time()
                row[5].value = "Early" if check_out_time < check_out_deadline else "Late"
                workbook.save(file_name)
                return f"Check-out time recorded as {row[5].value} for {name}."

            return f"{name} has already checked in today."

    if action == "check-in":
        # Determine status based on check-in time
        check_in_time = datetime.strptime(time_now, '%H:%M:%S').time()
        check_in_status = "Early" if check_in_time <= check_in_deadline else "Late"
        sheet.append([name, today, time_now, check_in_status, None, None])
        workbook.save(file_name)
        return f"Check-in time recorded as {check_in_status} for {name}."

    workbook.save(file_name)
    return f"No check-in record found for {name} to update check-out."
