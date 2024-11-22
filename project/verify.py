import cv2
import numpy as np
import face_recognition
import os
from openpyxl import Workbook, load_workbook
import attandance

# Function to load data from Excel
def load_data_from_excel(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    data_dict = {}
    images = []
    classNames = []
    imagePaths = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]  # Name is assumed to be in the first column
        info = {
            "name": name,  # Include name in info for display
            "age": row[1],
            "gender": row[2],
            "hometown": row[3],
            "address": row[4],
            "Image": row[5]  # Assuming the image path is in the 6th column
        }
        data_dict[name] = info
        curImg = cv2.imread(info["Image"])
        if curImg is not None:
            images.append(curImg)
            classNames.append(name)
            imagePaths.append(info["Image"])
        else:
            print(f"Warning: Failed to load image for {name} from path: {info['Image']}")
    
    return data_dict, images, classNames, imagePaths

# Function to find encodings of images
def find_encodings(images):
    encodeList = []
    for img in images:
        try:
            encodes = face_recognition.face_encodings(img)
            if encodes:  # Check if encodes are found to avoid empty lists
                encodeList.append(encodes[0])
            else:
                print("Warning: No encodings found for image.")
        except Exception as e:
            print(f"Error encoding image: {e}")
    return encodeList

# Function to process the face and return necessary information
def process_face(img, facesCurFrame, encodesCurFrame, encodeListKnown, classNames, data_dict, face_recognition_start_time):
    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        # Check if the face matches any known face encoding
        name, matches, faceDis = recognize_face(encodeFace, encodeListKnown, classNames)
        
        # Draw rectangle and name above the face
        y1, x2, y2, x1 = faceLoc
        y1, x2, y2, x1 = scale_face_locations(y1, x2, y2, x1)
        
        # Draw the rectangle and display the name
        rectangle_color = (0, 0, 255) if name == "Unknown" else (0, 255, 0)
        draw_face_rectangle(img, x1, y1, x2, y2, rectangle_color)
        display_face_name(img, x1, y1, name)
        
        # Fetch data for recognized faces or handle unknown
        if name != "Unknown":
            info = data_dict.get(name, {})
            if info:
                display_info_card(img, x1, y1, x2, info)
                track_and_mark_attendance(name, face_recognition_start_time)
        else:
            # Display "Unknown" info on the card
            display_info_card(img, x1, y1, x2, {"name": "Unknown", "age": "N/A", "gender": "N/A", "hometown": "N/A", "address": "N/A"})
    
    return img, face_recognition_start_time


# Recognize the face and return the name, matches, and distances
def recognize_face(encodeFace, encodeListKnown, classNames):
    matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
    faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
    name = "Unknown"  # Default to "Unknown" if no match is found
    if matches and any(matches):
        matchIndex = np.argmin(faceDis)
        if matches[matchIndex]:
            name = classNames[matchIndex]  # Update name if a match is found
    return name, matches, faceDis


# Scale the face locations from small image to original image size
def scale_face_locations(y1, x2, y2, x1):
    y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
    return y1, x2, y2, x1


# Draw rectangle around the face
def draw_face_rectangle(img, x1, y1, x2, y2, rectangle_color):
    cv2.rectangle(img, (x1, y1), (x2, y2), rectangle_color, 2)


# Display name above the rectangle
def display_face_name(img, x1, y1, name):
    label_y = y1 - 10 if y1 - 10 > 10 else y1 + 20
    cv2.putText(img, name, (x1, label_y), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)


# Function to display information card
def display_info_card(img, x1, y1, x2, info):
    card_x1, card_y1 = x1, y1 - 130
    card_x2, card_y2 = x2, y1 - 10

    # Draw card background
    cv2.rectangle(img, (card_x1, card_y1), (card_x2, card_y2), (0, 0, 255), cv2.FILLED)
    
    # Draw connecting line
    cv2.line(img, (x1, y1), (card_x1, card_y2), (255, 255, 255), 2)

    # Add text to the card
    cv2.putText(img, f"Name: {info.get('name', 'N/A')}", (card_x1 + 6, card_y1 + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
    cv2.putText(img, f"Age: {info.get('age', 'N/A')}", (card_x1 + 6, card_y1 + 40), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
    cv2.putText(img, f"Gender: {info.get('gender', 'N/A')}", (card_x1 + 6, card_y1 + 60), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
    cv2.putText(img, f"Hometown: {info.get('hometown', 'N/A')}", (card_x1 + 6, card_y1 + 80), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
    cv2.putText(img, f"Class: {info.get('address', 'N/A')}", (card_x1 + 6, card_y1 + 100), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)

# Function to track and mark attendance after 8 seconds
def track_and_mark_attendance(name, face_recognition_start_time):
    current_time = cv2.getTickCount() / cv2.getTickFrequency()

    if name not in face_recognition_start_time:
        face_recognition_start_time[name] = current_time  # Start tracking time
    elif current_time - face_recognition_start_time[name] >= 8:
        # Mark attendance after 8 seconds
        attandance.markAttendance(name)
        print(f"Attendance marked for {name}.")
        del face_recognition_start_time[name]  # Reset timer for this person

# Main function to verify faces and mark attendance
def verifydata():
    cap = cv2.VideoCapture(0)
    face_recognition_start_time = {}  # Initialize the tracking dictionary

    # Load data from Excel and find encodings
    excel_file = "data.xlsx"
    data_dict, images, classNames, imagePaths = load_data_from_excel(excel_file)
    encodeListKnown = find_encodings(images)

    while True:
        success, img = cap.read()
        if not success:
            print("Failed to grab frame from webcam.")
            break

        imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

        facesCurFrame = face_recognition.face_locations(imgS)
        encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

        # Process faces and mark attendance if necessary
        img, face_recognition_start_time = process_face(img, facesCurFrame, encodesCurFrame, encodeListKnown, classNames, data_dict, face_recognition_start_time)

        cv2.imshow('Webcam', img)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

