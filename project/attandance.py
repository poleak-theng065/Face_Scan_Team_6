from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# Function to check attendance
def markAttendance(name):
    file_name = 'Attendance.xlsx'
    
    # Check if the Excel file exists, if not, create it and add the header
    if not os.path.isfile(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Attendance'
        sheet.append(['Name', 'Date', 'Time', 'DateTime'])  # Adding the header
        workbook.save(file_name)
    
    # Open the Excel file and access the active sheet
    workbook = load_workbook(file_name)
    sheet = workbook.active
    
    # Collect existing names to prevent duplicates
    nameList = [sheet.cell(row=row, column=1).value for row in range(2, sheet.max_row + 1)]
    
    # If the name is not already in the sheet, add it with date, time, and dateTime
    if name not in nameList:
        now = datetime.now()
        dateString = now.strftime('%Y-%m-%d')       # Current date
        timeString = now.strftime('%H:%M:%S')       # Current time
        dateTimeString = now.strftime('%Y-%m-%d %H:%M:%S')  # Full dateTime
        
        # Append name, date, time, and dateTime to the sheet
        sheet.append([name, dateString, timeString, dateTimeString])
        workbook.save(file_name)  # Save changes to the file
        print(f"Attendance marked for {name}.")
    else:
        print(f"{name} is already marked as present.")


