import cv2
import numpy as np
import face_recognition
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import pyttsx3  # For text-to-speech

# Initialize the text-to-speech engine
engine = pyttsx3.init()

def speak_greeting(name):
    """
    Generate a greeting based on the current time and name.
    """
    current_hour = datetime.now().hour
    if current_hour < 12:
        greeting = f"Good morning {name}! Have a productive day."
    elif current_hour < 18:
        greeting = f"Good afternoon {name}! Hope your day is going well."
    else:
        greeting = f"Good evening {name}! Great to see you."
    engine.say(greeting)
    engine.runAndWait()

def markAttendance(name, action):
    """
    Marks attendance for a given name and action (check-in or check-out).
    """
    file_name = 'CheckAttendance.xlsx'
    
    # Create file if it doesn't exist
    if not os.path.isfile(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Name', 'Date', 'Check-In Time', 'Check-In Status', 'Check-Out Time', 'Check-Out Status'])
        workbook.save(file_name)

    workbook = load_workbook(file_name)
    sheet = workbook.active
    now = datetime.now()
    today_date = now.strftime('%Y-%m-%d')
    current_time = now.strftime('%H:%M:%S')

    check_in_deadline = datetime.strptime('07:00:00', '%H:%M:%S').time()
    check_out_deadline = datetime.strptime('16:00:00', '%H:%M:%S').time()

    for row in sheet.iter_rows(min_row=2, values_only=False):
        print(f"Checking row: {row[0].value}, Date: {row[1].value}")  # Debug print

        if row[0].value == name and row[1].value == today_date:
            print(f"Found existing record for {name} on {today_date}")  # Debug print
            
            if action == "check-out":
                if row[2].value is None:
                    print(f"{name} has not checked in today. Check-in needed first.")  # Debug print
                    return f"{name} hasn't checked in today. Please check-in first."
                
                if row[4].value is None:
                    row[4].value = current_time
                    check_out_time = datetime.strptime(current_time, '%H:%M:%S').time()
                    row[5].value = "You are working on time." if check_out_time < check_out_deadline else "You are working late."
                    workbook.save(file_name)
                    speak_greeting(name)
                    return f"Check-out time recorded as {row[5].value} for {name}."
                else:
                    return f"{name} has already checked out today."

            elif action == "check-in":
                if row[2].value is None:  # Check if already checked in
                    check_in_time = datetime.strptime(current_time, '%H:%M:%S').time()
                    check_in_status = "You are leaving after hours." if check_in_time <= check_in_deadline else "You are overtime."
                    row[2].value = current_time
                    row[3].value = check_in_status
                    workbook.save(file_name)
                    speak_greeting(name)
                    return f"Check-in time recorded as {check_in_status} for {name}."
                else:
                    return f"{name} has already checked in today."

    # If no record exists for the name and date, create a new check-in record
    if action == "check-in":
        print(f"No existing record for {name} today, creating new check-in.")  # Debug print
        check_in_time = datetime.strptime(current_time, '%H:%M:%S').time()
        check_in_status = "You are leaving after hours." if check_in_time <= check_in_deadline else "You are overtime."
        sheet.append([name, today_date, current_time, check_in_status, None, None])
        workbook.save(file_name)
        speak_greeting(name)
        return f"Check-in time recorded as {check_in_status} for {name}."

    return f"No check-in record found for {name} to update check-out."


def load_data_from_excel(excel_file):
    """
    Load student data from an Excel file and return images, names, and info dictionary.
    """
    wb = load_workbook(excel_file)
    ws = wb.active

    data_dict = {}
    images = []
    classNames = []
    imagePaths = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        info = {
            "name": name,
            "age": row[1],
            "gender": row[2],
            "hometown": row[3],
            "address": row[4],
            "Image": row[5]
        }
        data_dict[name] = info
        curImg = cv2.imread(info["Image"])
        if curImg is not None:
            images.append(curImg)
            classNames.append(name)
            imagePaths.append(info["Image"])
        else:
            print(f"Warning: Failed to load image for {name} from path: {info['Image']}")

    return data_dict, images, classNames, imagePaths

def find_encodings(images):
    """
    Find encodings for the given images.
    """
    encodeList = []
    for img in images:
        try:
            encodes = face_recognition.face_encodings(img)
            if encodes:
                encodeList.append(encodes[0])
            else:
                print("Warning: No encodings found for image.")
        except Exception as e:
            print(f"Error encoding image: {e}")
    return encodeList

def display_info_card_with_animation(img, x1, y1, x2, info, frame_count):
    """
    Display an animated information card on the frame.
    """
    card_x1, card_y1 = x1, y1 - 130
    card_x2, card_y2 = x2, y1 - 10
    overlay = img.copy()
    alpha = 0.2
    cv2.rectangle(overlay, (card_x1, card_y1), (card_x2, card_y2), cv2.FILLED)
    cv2.addWeighted(overlay, alpha, img, 1 - alpha, 0, img)
    cv2.line(img, (x1, y1), (card_x1, card_y2), (255, 255, 255), 2)

    keys = ["name", "age", "gender", "hometown", "address"]
    labels = ["Name", "Age", "Gender", "Hometown", "Class"]
    max_index = min(frame_count // 2, len(keys))

    for i in range(max_index):
        cv2.putText(
            img,
            f"{labels[i]}: {info.get(keys[i], 'N/A')}",
            (card_x1 + 6, card_y1 + 20 + i * 20),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.5,
            (255, 255, 255),
            1
        )

def track_and_mark_attendance(name, face_recognition_start_time, action):
    """
    Tracks the duration a face is recognized and marks attendance if time threshold is met.
    """
    current_time = cv2.getTickCount() / cv2.getTickFrequency()

    if name not in face_recognition_start_time:
        face_recognition_start_time[name] = current_time
    elif current_time - face_recognition_start_time[name] >= 8:
        result = markAttendance(name, action)
        print(f"Attendance marked for {name}: {result}")
        del face_recognition_start_time[name]

def process_face(img, facesCurFrame, encodesCurFrame, encodeListKnown, classNames, data_dict, face_recognition_start_time, frame_count_dict, action):
    """
    Process faces in the frame, identify, and display info or mark attendance.
    """
    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)

        name = "Unknown"
        if matches and any(matches):
            matchIndex = np.argmin(faceDis)
            if matches[matchIndex]:
                name = classNames[matchIndex]

        y1, x2, y2, x1 = faceLoc
        y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
        rectangle_color = (0, 0, 255) if name == "Unknown" else (0, 255, 0)
        cv2.rectangle(img, (x1, y1), (x2, y2), rectangle_color, 2)

        if name not in frame_count_dict:
            frame_count_dict[name] = 0

        if name != "Unknown":
            info = data_dict.get(name, {})
            if info:
                display_info_card_with_animation(img, x1, y1, x2, info, frame_count_dict[name])
                track_and_mark_attendance(name, face_recognition_start_time, action)
        else:
            display_info_card_with_animation(img, x1, y1, x2, {
                "name": "Unknown",
                "age": "N/A",
                "gender": "N/A",
                "hometown": "N/A",
                "address": "N/A"
            }, frame_count_dict[name])

        frame_count_dict[name] += 1

    return img, face_recognition_start_time, frame_count_dict

def verifydata(action):
    """
    Main function to capture webcam feed, detect faces, and process attendance.
    """
    cap = cv2.VideoCapture(0)
    face_recognition_start_time = {}
    frame_count_dict = {}

    excel_file = "dataOfStudent.xlsx"
    data_dict, images, classNames, imagePaths = load_data_from_excel(excel_file)
    encodeListKnown = find_encodings(images)

    while True:
        success, img = cap.read()
        if not success:
            print("Failed to grab frame from webcam.")
            break

        imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

        facesCurFrame = face_recognition.face_locations(imgS)
        encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

        img, face_recognition_start_time, frame_count_dict = process_face(
            img, facesCurFrame, encodesCurFrame, encodeListKnown, classNames,
            data_dict, face_recognition_start_time, frame_count_dict, action
        )

        cv2.imshow('Face Recognition Attendance', img)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

# Example usage
verifydata("check-out")
