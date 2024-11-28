import cv2
import numpy as np
import face_recognition
from openpyxl import  load_workbook
from tkinter import filedialog, messagebox
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
import pyttsx3  # For text-to-speech

# Initialize the text-to-speech engine
engine = pyttsx3.init()

# Function to speak a greeting
def speak_greeting(name):
    current_hour = datetime.now().hour
    if current_hour < 12:
        greeting = f"Good morning {name}! Have a productive day."
    elif current_hour < 18:
        greeting = f"Good afternoon {name}! Hope your day is going well."
    else:
        greeting = f"Good evening {name}! Great to see you."
    engine.say(greeting)
    engine.runAndWait()

# Function to manage attendance (check-in and check-out)
def markAttendance(name, action="check"):
    file_name = 'CheckAttendance.xlsx'
    
    # Check if the file exists. If not, create it and add headers
    if not os.path.isfile(file_name):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Name', 'Date', 'Check-In Time', 'Check-In Status', 'Check-Out Time', 'Check-Out Status'])
        workbook.save(file_name)
    
    # Load the existing workbook
    workbook = load_workbook(file_name)
    sheet = workbook.active
    now = datetime.now()
    today_date = now.strftime('%Y-%m-%d')
    current_time = now.strftime('%H:%M:%S')

    # Define thresholds for check-in and check-out
    check_in_deadline = datetime.strptime('07:00:00', '%H:%M:%S').time()
    check_out_deadline = datetime.strptime('16:00:00', '%H:%M:%S').time()

    # Loop through all rows in the sheet to see if the student is already in the attendance list for today
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == name and row[1].value == today_date:
            if action == "check-out" and row[4].value is None:  # If check-out is not already marked
                row[4].value = current_time
                check_out_time = datetime.strptime(current_time, '%H:%M:%S').time()
                row[5].value = "Early" if check_out_time < check_out_deadline else "Late"
                workbook.save(file_name)
                return f"Check-out time recorded as {row[5].value} for {name}."
            return f"{name} has already checked in today."

    # If no existing check-in record found, add a new entry for check-in
    if action == "check-in":
        check_in_time = datetime.strptime(current_time, '%H:%M:%S').time()
        check_in_status = "Early" if check_in_time <= check_in_deadline else "Late"
        sheet.append([name, today_date, current_time, check_in_status, None, None])
        workbook.save(file_name)
        return f"Check-in time recorded as {check_in_status} for {name}."

    return f"No check-in record found for {name} to update check-out."



# Function to load data from Excel
def load_data_from_excel(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    data_dict = {}
    images = []
    classNames = []
    imagePaths = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]  # Name is assumed to be in the first column
        info = {
            "name": name,  # Include name in info for display
            "age": row[1],
            "gender": row[2],
            "hometown": row[3],
            "address": row[4],
            "Image": row[5]  # Assuming the image path is in the 6th column
        }
        data_dict[name] = info
        curImg = cv2.imread(info["Image"])
        if curImg is not None:
            images.append(curImg)
            classNames.append(name)
            imagePaths.append(info["Image"])
        else:
            print(f"Warning: Failed to load image for {name} from path: {info['Image']}")
    
    return data_dict, images, classNames, imagePaths

# Function to find encodings of images
def find_encodings(images):
    encodeList = []
    for img in images:
        try:
            encodes = face_recognition.face_encodings(img)
            if encodes:  # Check if encodes are found to avoid empty lists
                encodeList.append(encodes[0])
            else:
                print("Warning: No encodings found for image.")
        except Exception as e:
            print(f"Error encoding image: {e}")
    return encodeList

# Function to process the face and return necessary information
def process_face(img, facesCurFrame, encodesCurFrame, encodeListKnown, classNames, data_dict, face_recognition_start_time, frame_count_dict):
    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)

        name = "Unknown"  # Default to "Unknown" if no match is found
        if matches and any(matches):
            matchIndex = np.argmin(faceDis)
            if matches[matchIndex]:
                name = classNames[matchIndex]  # Update name if a match is found

        # Draw rectangle around the face
        y1, x2, y2, x1 = faceLoc
        y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4

        # Determine rectangle color based on recognition
        rectangle_color = (0, 0, 255) if name == "Unknown" else (0, 255, 0)
        cv2.rectangle(img, (x1, y1), (x2, y2), rectangle_color, 2)

        # Initialize frame count for new faces
        if name not in frame_count_dict:
            frame_count_dict[name] = 0

        # Fetch data for recognized faces or display "Unknown" info
        if name != "Unknown":
            info = data_dict.get(name, {})
            if info:
                display_info_card_with_animation(img, x1, y1, x2, info, frame_count_dict[name])
                track_and_mark_attendance(name, face_recognition_start_time)
        else:
            # Display "Unknown" info on the card
            display_info_card_with_animation(img, x1, y1, x2, {
                "name": "Unknown",
                "age": "N/A",
                "gender": "N/A",
                "hometown": "N/A",
                "address": "N/A"
            }, frame_count_dict[name])

        # Increment frame count for this face
        frame_count_dict[name] += 1

    return img, face_recognition_start_time, frame_count_dict


# Function to display information card with animation
def display_info_card_with_animation(img, x1, y1, x2, info, frame_count):
    card_x1, card_y1 = x1, y1 - 130
    card_x2, card_y2 = x2, y1 - 10

    # Create an overlay for transparency
    overlay = img.copy()

    # Draw a transparent rectangle for the card
    alpha = 0.3  # Transparency factor
    cv2.rectangle(overlay, (card_x1, card_y1), (card_x2, card_y2), cv2.FILLED)
    cv2.addWeighted(overlay, alpha, img, 1 - alpha, 0, img)

    # Draw the connecting line
    cv2.line(img, (x1, y1), (card_x1, card_y2), (255, 255, 255), 2)

    # Define keys and labels for information
    keys = ["name", "age", "gender", "hometown", "address"]
    labels = ["Name", "Age", "Gender", "Hometown", "Class"]
    
    # Determine the number of lines to display based on frame count
    max_index = min(frame_count // 2, len(keys))  # Show one line every 10 frames

    # Add text incrementally
    for i in range(max_index):
        cv2.putText(
            img,
            f"{labels[i]}: {info.get(keys[i], 'N/A')}",
            (card_x1 + 6, card_y1 + 20 + i * 20),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.5,
            (255, 255, 255),
            1
        )

# Function to track and mark attendance after 8 seconds
def track_and_mark_attendance(name, face_recognition_start_time):
    current_time = cv2.getTickCount() / cv2.getTickFrequency()

    if name not in face_recognition_start_time:
        face_recognition_start_time[name] = current_time  # Start tracking time
    elif current_time - face_recognition_start_time[name] >= 8:
        # Mark attendance after 8 seconds
        markAttendance(name,action="check")
        print(f"Attendance marked for {name} has chekout sucessfully.")
        del face_recognition_start_time[name]  # Reset timer for this person

# Main function to verify faces and mark attendance
def verifydata_checkin():
    cap = cv2.VideoCapture(0)
    face_recognition_start_time = {}  # Initialize the tracking dictionary
    frame_count_dict = {}  # Initialize the frame count dictionary for animations

    # Load data from Excel and find encodings
    excel_file = "dataOfStudent.xlsx"
    data_dict, images, classNames, imagePaths = load_data_from_excel(excel_file)
    encodeListKnown = find_encodings(images)

    while True:
        success, img = cap.read()
        if not success:
            print("Failed to grab frame from webcam.")
            break

        imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

        facesCurFrame = face_recognition.face_locations(imgS)
        encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

        # Process faces and mark attendance if necessary
        img, face_recognition_start_time, frame_count_dict = process_face(
            img, facesCurFrame, encodesCurFrame, encodeListKnown, classNames, data_dict, face_recognition_start_time, frame_count_dict
        )

        cv2.imshow('Webcam', img)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()